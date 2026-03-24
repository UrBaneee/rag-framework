"""XLSX parser using openpyxl — extracts sheet rows as text blocks."""

import hashlib
import logging
from pathlib import Path

from rag.core.contracts.document import Document
from rag.core.contracts.ir_block import BlockType, IRBlock
from rag.core.contracts.parse_report import ParseReport
from rag.core.interfaces.parser import BaseParser

logger = logging.getLogger(__name__)

_PARSER_NAME = "xlsx"
_SUPPORTED_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_MIN_BLOCK_CHARS = 1

try:
    import openpyxl as _openpyxl

    _XLSX_AVAILABLE = True
except ImportError:  # pragma: no cover
    _XLSX_AVAILABLE = False


class XlsxParser(BaseParser):
    """Parser for .xlsx files using openpyxl.

    Each non-empty row across all sheets is emitted as a PARAGRAPH IRBlock.
    The sheet name is prepended to the first row of each sheet as a HEADING
    block so downstream chunkers can use it as context.

    Raises:
        ImportError: If openpyxl is not installed (raised at instantiation).

    Usage::

        parser = XlsxParser()
        doc = parser.parse("/path/to/file.xlsx")
    """

    def __init__(self) -> None:
        if not _XLSX_AVAILABLE:
            raise ImportError(
                "openpyxl is not installed. Install with: pip install openpyxl"
            )

    def supports(self, mime_type: str) -> bool:
        return mime_type == _SUPPORTED_MIME_TYPE

    def parse(self, source_path: str) -> Document:
        """Parse an .xlsx file into a Document.

        Args:
            source_path: Absolute path to the .xlsx file.

        Returns:
            Document with IRBlocks — one HEADING per sheet, one PARAGRAPH
            per non-empty row.

        Raises:
            ValueError: If the file cannot be opened.
        """
        path = Path(source_path)
        try:
            wb = _openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open XLSX '{source_path}': {exc}") from exc

        blocks: list[IRBlock] = []

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Emit sheet name as a heading
                blocks.append(
                    IRBlock(block_type=BlockType.HEADING, text=f"Sheet: {sheet_name}")
                )

                for row in ws.iter_rows(values_only=True):
                    cell_texts = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if not cell_texts:
                        continue
                    row_text = " | ".join(cell_texts)
                    if len(row_text) >= _MIN_BLOCK_CHARS:
                        blocks.append(
                            IRBlock(block_type=BlockType.PARAGRAPH, text=row_text)
                        )
        finally:
            wb.close()

        all_text = " ".join(b.text for b in blocks)
        report = ParseReport(
            char_count=len(all_text),
            block_count=len(blocks),
            non_printable_ratio=0.0,
            repetition_score=0.0,
            parser_used=_PARSER_NAME,
            fallback_triggered=False,
        )

        doc_id = hashlib.sha256(source_path.encode()).hexdigest()[:16]
        return Document(
            doc_id=doc_id,
            source_path=source_path,
            mime_type=_SUPPORTED_MIME_TYPE,
            metadata={"filename": path.name, "extension": "xlsx"},
            blocks=blocks,
            parse_report=report,
        )
